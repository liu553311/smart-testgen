"""Excel (.xlsx) exporter using openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from smart_testgen.core.models import TestSuite
from smart_testgen.exporters.base import BaseExporter

# Priority-based row colors
PRIORITY_COLORS = {
    "critical": "FF4444",  # red
    "high": "FF8C00",      # orange
    "medium": "FFD700",    # yellow
    "low": "90EE90",       # light green
}

HEADERS = [
    "ID",
    "Title",
    "Description",
    "Priority",
    "Category",
    "Preconditions",
    "Steps",
    "Expected Results",
    "Tags",
]


class ExcelExporter(BaseExporter):
    """Export test cases as a styled Excel workbook."""

    def export(self, test_suite: TestSuite, output_path: Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Header row
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, tc in enumerate(test_suite.test_cases, 2):
            steps_text = "\n".join(
                f"{s.step_number}. {s.action} → {s.expected_result}" for s in tc.steps
            )

            values = [
                tc.id,
                tc.title,
                tc.description,
                tc.priority.value,
                tc.category.value,
                "\n".join(tc.preconditions) if tc.preconditions else "",
                steps_text,
                "\n".join(tc.expected_results),
                ", ".join(tc.tags) if tc.tags else "",
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Apply priority-based row color
            color = PRIORITY_COLORS.get(tc.priority.value)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Auto-adjust column widths (approximate)
        col_widths = {
            1: 10,   # ID
            2: 30,   # Title
            3: 50,   # Description
            4: 12,   # Priority
            5: 14,   # Category
            6: 30,   # Preconditions
            7: 50,   # Steps
            8: 40,   # Expected Results
            9: 20,   # Tags
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(str(output_path))
        return output_path
