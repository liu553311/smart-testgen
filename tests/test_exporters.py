"""Tests for exporters (Markdown, JSON, Excel)."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from smart_testgen.core.models import (
    TestCase,
    TestCategory,
    TestPriority,
    TestStep,
    TestSuite,
)
from smart_testgen.exporters.markdown import MarkdownExporter
from smart_testgen.exporters.json_exporter import JsonExporter
from smart_testgen.exporters.excel import ExcelExporter


@pytest.fixture
def sample_suite() -> TestSuite:
    return TestSuite(
        title="Test Suite",
        source_requirements="Short requirements text",
        generated_at="2026-06-10T12:00:00+00:00",
        provider_used="openai",
        model_used="gpt-4o",
        test_cases=[
            TestCase(
                id="TC-001",
                title="Login test",
                description="Verify login works",
                priority=TestPriority.HIGH,
                category=TestCategory.FUNCTIONAL,
                preconditions=["User exists"],
                steps=[
                    TestStep(step_number=1, action="Go to login", expected_result="Page loads"),
                    TestStep(step_number=2, action="Enter creds", expected_result="Accepted"),
                ],
                expected_results=["User logged in"],
                tags=["login"],
            ),
            TestCase(
                id="TC-002",
                title="Boundary email",
                description="Test max email length",
                priority=TestPriority.MEDIUM,
                category=TestCategory.BOUNDARY,
                preconditions=[],
                steps=[
                    TestStep(step_number=1, action="Enter 254-char email", expected_result="Accepted"),
                ],
                expected_results=["No truncation"],
                tags=["email", "boundary"],
            ),
        ],
    )


class TestMarkdownExporter:
    def test_export_creates_file(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = MarkdownExporter()
        out = exporter.export(sample_suite, tmp_path / "output.md")
        assert out.exists()
        assert out.suffix == ".md"

    def test_export_contains_headers(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = MarkdownExporter()
        out = exporter.export(sample_suite, tmp_path / "output.md")
        content = out.read_text(encoding="utf-8")
        assert "# Test Suite" in content
        assert "TC-001" in content
        assert "TC-002" in content
        assert "Login test" in content

    def test_export_contains_summary(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = MarkdownExporter()
        out = exporter.export(sample_suite, tmp_path / "output.md")
        content = out.read_text(encoding="utf-8")
        assert "Summary by Category" in content
        assert "functional" in content
        assert "boundary" in content

    def test_creates_parent_dirs(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = MarkdownExporter()
        out = exporter.export(sample_suite, tmp_path / "sub" / "dir" / "output.md")
        assert out.exists()


class TestJsonExporter:
    def test_export_creates_file(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = JsonExporter()
        out = exporter.export(sample_suite, tmp_path / "output.json")
        assert out.exists()

    def test_export_valid_json(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = JsonExporter()
        out = exporter.export(sample_suite, tmp_path / "output.json")
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["title"] == "Test Suite"
        assert len(data["test_cases"]) == 2

    def test_export_preserves_fields(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = JsonExporter()
        out = exporter.export(sample_suite, tmp_path / "output.json")
        data = json.loads(out.read_text(encoding="utf-8"))
        tc = data["test_cases"][0]
        assert tc["id"] == "TC-001"
        assert tc["priority"] == "high"
        assert tc["category"] == "functional"


class TestExcelExporter:
    def test_export_creates_file(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = ExcelExporter()
        out = exporter.export(sample_suite, tmp_path / "output.xlsx")
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_export_headers(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = ExcelExporter()
        out = exporter.export(sample_suite, tmp_path / "output.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers[0] == "ID"
        assert headers[1] == "Title"
        assert headers[3] == "Priority"

    def test_export_data_rows(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = ExcelExporter()
        out = exporter.export(sample_suite, tmp_path / "output.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        # Row 2 = first test case
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=2, column=2).value == "Login test"
        # Row 3 = second test case
        assert ws.cell(row=3, column=1).value == "TC-002"

    def test_export_multiple_cases(self, tmp_path: Path, sample_suite: TestSuite) -> None:
        exporter = ExcelExporter()
        out = exporter.export(sample_suite, tmp_path / "output.xlsx")
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 3  # header + 2 data rows
